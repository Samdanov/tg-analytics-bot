# app/services/usecases/website_service.py

"""
Сервис для анализа веб-сайтов.
Аналогичен channel_service, но работает с сайтами.
"""

from pathlib import Path
from typing import Optional
from datetime import datetime

from app.services.website_parser.parser import parse_website
from app.services.llm.analyzer import analyze_text_content
from app.services.similarity_engine.website_similarity import find_similar_channels_by_keywords
from app.services.xlsx_generator import generate_similar_channels_xlsx
from app.core.logging import get_logger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = get_logger(__name__)


def _auto_width(ws):
    """Автоматически подгоняет ширину колонок."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except Exception:
                value = ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


async def generate_website_similar_channels_xlsx(
    url: str,
    similar_channels: list,
    keywords: list,
    base_dir: Optional[Path] = None
) -> Path:
    """
    Генерирует Excel файл с похожими каналами для сайта.
    
    Args:
        url: URL сайта
        similar_channels: Список похожих каналов
        keywords: Ключевые слова из анализа
        base_dir: Базовая директория для сохранения
    
    Returns:
        Путь к созданному файлу
    """
    if base_dir is None:
        base_dir = Path(__file__).resolve().parents[3]
    
    reports_dir = base_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Создаем безопасное имя файла
    safe_url = url.replace("https://", "").replace("http://", "").replace("/", "_")[:50]
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = reports_dir / f"similar_channels_website_{safe_url}_{ts}.xlsx"
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Каналы"
    
    # Заголовок
    ws.append([
        "Дата создания",
        "Релевантность, %",
        "Имя канала",
        "Подписчики",
        "Название канала",
        "Описание канала",
    ])
    
    created_str = datetime.utcnow().strftime("%d.%m.%Y")
    
    # Находим максимальный score для нормализации
    max_score = 0.0
    if similar_channels:
        scores = [ch.get("score", 0.0) for ch in similar_channels]
        max_score = max(scores) if scores else 1.0
    
    # Заполняем данные
    for ch_data in similar_channels:
        score = float(ch_data.get("score", 0.0))
        
        # Нормализуем score до 0-100%
        if max_score > 0:
            normalized_score = min(score / max_score, 1.0)
        else:
            normalized_score = min(score, 1.0)
        
        relevance_percent = round(normalized_score * 100, 1)
        
        username = ch_data.get("username", "")
        if username.startswith("id:"):
            link = ""
        else:
            link = f"https://t.me/{username}" if username else ""
        
        ws.append([
            created_str,
            relevance_percent,
            link,
            ch_data.get("subscribers"),
            ch_data.get("title"),
            "",  # Описание не доступно для сайтов
        ])
    
    _auto_width(ws)
    wb.save(filename)
    
    logger.info(f"Generated Excel report: {filename}")
    return filename


async def run_website_analysis_pipeline(url: str, top_n: int = 10) -> tuple[Path, dict]:
    """
    Полный пайплайн анализа веб-сайта:
    1. Парсинг сайта
    2. Анализ через LLM (ключевые слова, ЦА)
    3. Поиск похожих каналов по ключевым словам
    4. Генерация Excel отчета
    
    Args:
        url: URL сайта
        top_n: Количество похожих каналов
    
    Returns:
        Путь к Excel файлу
    """
    logger.info(f"[WEBSITE] Начало анализа сайта: {url}")
    
    # 1. Парсинг сайта
    text, error = await parse_website(url)
    if error:
        logger.error(f"[WEBSITE] Ошибка парсинга: {error}")
        raise ValueError(f"Не удалось распарсить сайт: {error}")
    
    if not text:
        raise ValueError("Не удалось извлечь текст с сайта")
    
    logger.info(f"[WEBSITE] Извлечено текста: {len(text)} символов")
    
    # 2. Анализ через LLM
    logger.info(f"[WEBSITE] Запуск LLM анализа...")
    analysis_result = await analyze_text_content(
        text=text,
        title=url,
        description=""
    )
    
    keywords = analysis_result.get("keywords", [])
    audience = analysis_result.get("audience", "")
    logger.info(
        f"[WEBSITE] Анализ завершен. "
        f"Ключевые слова: {len(keywords)}, "
        f"ЦА: {len(audience)} символов, "
        f"Ключевые слова: {keywords[:5]}..."
    )
    
    if not keywords:
        raise ValueError("Не удалось извлечь ключевые слова из сайта")
    
    # 3. Поиск похожих каналов по ключевым словам
    similar_channels = await find_similar_channels_by_keywords(
        keywords=keywords,
        top_n=top_n
    )
    
    if not similar_channels:
        raise ValueError("Не найдено похожих каналов по ключевым словам")
    
    logger.info(f"[WEBSITE] Найдено похожих каналов: {len(similar_channels)}")
    
    # 4. Генерация Excel отчета
    report_path = await generate_website_similar_channels_xlsx(
        url=url,
        similar_channels=similar_channels,
        keywords=keywords
    )
    
    logger.info(f"[WEBSITE] Pipeline завершен. Отчет: {report_path}")
    return report_path, analysis_result
