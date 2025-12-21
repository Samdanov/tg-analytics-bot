# app/services/xlsx_generator.py

import json
import math
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER
from sqlalchemy import select
import re

from app.db.database import async_session_maker
from app.db.models import Channel, AnalyticsResults


def _auto_width(ws):
    """Автоматически подгоняем ширину колонок под содержимое."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except Exception:
                value = ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def _sanitize_filename(text: str, max_length: int = 50) -> str:
    """Очищает текст для использования в имени файла."""
    # Удаляем недопустимые символы для Windows/Linux
    text = re.sub(r'[<>:"/\\|?*]', '', text)
    # Заменяем пробелы на подчёркивания
    text = text.replace(' ', '_')
    # Удаляем множественные подчёркивания
    text = re.sub(r'_+', '_', text)
    # Обрезаем до максимальной длины
    if len(text) > max_length:
        text = text[:max_length]
    return text.strip('_')


def _apply_header_style(ws, header_row: int = 1):
    """Применяет стили к заголовкам таблицы."""
    # Стили для заголовков
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="000000")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Устанавливаем высоту строки заголовка
    ws.row_dimensions[header_row].height = 30


def _apply_data_style(ws, start_row: int = 2):
    """Применяет стили к данным таблицы."""
    border_side = Side(style="thin", color="CCCCCC")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Выравнивание для разных колонок
    alignments = {
        1: Alignment(horizontal="center", vertical="center"),  # № (номер)
        2: Alignment(horizontal="center", vertical="center"),  # Степень схожести
        3: Alignment(horizontal="left", vertical="top"),   # Ссылка
        4: Alignment(horizontal="center", vertical="top"),   # Подписчики
        5: Alignment(horizontal="left", vertical="top", wrap_text=True),  # Название
        6: Alignment(horizontal="left", vertical="top", wrap_text=True),  # Описание
    }
    
    # Цветовая палитра для степени схожести
    similarity_colors = {
        "Очень высокая": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),  # Зеленый
        "Высокая": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),        # Светло-зеленый
        "Средняя": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),        # Оранжевый
        "Низкая": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),         # Красный
    }
    
    for row in ws.iter_rows(min_row=start_row):
        for idx, cell in enumerate(row, start=1):
            cell.border = data_border
            if idx in alignments:
                cell.alignment = alignments[idx]
            
            # Форматирование для степени схожести (колонка 2)
            if idx == 2 and isinstance(cell.value, str):
                # Применяем цвет в зависимости от текста
                if cell.value in similarity_colors:
                    cell.fill = similarity_colors[cell.value]
                    cell.font = Font(bold=True, color="FFFFFF", size=11)  # Белый жирный текст
            
            # Форматирование для подписчиков (колонка 4)
            if idx == 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            
            # Форматирование ссылок (колонка 3)
            if idx == 3 and cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.font = Font(color="0563C1", underline="single")
                cell.hyperlink = cell.value


async def _load_source_channel(channel_username: str) -> Optional[Channel]:
    """Находим канал по username."""
    async with async_session_maker() as session:
        q = select(Channel).where(Channel.username == channel_username)
        res = await session.execute(q)
        return res.scalar_one_or_none()


async def _load_similar_results(source_channel_id: int) -> tuple[Optional[AnalyticsResults], List[int]]:
    """
    Берём последний AnalyticsResults для канала
    и список id похожих каналов в нужном порядке.
    """
    async with async_session_maker() as session:
        q = (
            select(AnalyticsResults)
            .where(AnalyticsResults.channel_id == source_channel_id)
            .order_by(AnalyticsResults.created_at.desc())
            .limit(1)
        )
        res = await session.execute(q)
        row: Optional[AnalyticsResults] = res.scalar_one_or_none()

    # Если записи нет → вернём "нет результата" (но НЕ ошибка!)
    if not row:
        return None, []

    # Если similar_channels_json пустой → вернём пустой список (НО НЕ ОШИБКА)
    try:
        raw = json.loads(row.similar_channels_json or "[]")
        if isinstance(raw, list):
            ids = [int(item.get("channel_id")) for item in raw if "channel_id" in item]
        else:
            ids = []
    except Exception:
        ids = []

    return row, ids


async def _load_channels_by_ids(ids: List[int]) -> Dict[int, Channel]:
    """Берём из БД все каналы с нужными id."""
    if not ids:
        return {}

    async with async_session_maker() as session:
        q = select(Channel).where(Channel.id.in_(ids))
        res = await session.execute(q)
        rows: List[Channel] = res.scalars().all()

    return {c.id: c for c in rows}


async def generate_similar_channels_xlsx(
    source_username: str,
    base_dir: Optional[Path] = None,
    top_n: int = 500,
) -> Path:
    """
    Генерирует XLSX-файл в формате skladtech_expo_*.
    Работает даже если похожих каналов нет.
    
    Args:
        source_username: username канала (с @ или без)
        base_dir: базовая директория для сохранения отчета
        top_n: максимальное количество похожих каналов в отчете (1-500)
    """

    # 1) Находим исходный канал
    if source_username.startswith("@"):
        source_username = source_username[1:]

    source_channel = await _load_source_channel(source_username)
    if not source_channel:
        raise ValueError(f"Канал @{source_username} не найден в БД")

    # 2) Берём последний результат similarity
    last_result, similar_ids = await _load_similar_results(source_channel.id)

    # ⚠️ Если результата similarity нет — создаём пустой XLSX, но НЕ ошибка
    if not last_result:
        similar_list = []
        channels_map = {}
    else:
        # 3) Загружаем каналы
        channels_map = await _load_channels_by_ids(similar_ids)

        # 4) Разбираем JSON
        try:
            # Используем глобальный json модуль, не локальный json_module
            raw = json.loads(last_result.similar_channels_json or "[]")
            similar_list = raw if isinstance(raw, list) else []
        except Exception as e:
            similar_list = []

    # 5) Готовим Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Похожие каналы"

    # Заголовки таблицы
    headers = [
        "№",
        "Степень схожести",
        "Ссылка на канал",
        "Подписчики",
        "Название канала",
        "Описание канала",
    ]
    ws.append(headers)
    
    # Применяем стили к заголовкам
    _apply_header_style(ws, header_row=1)


    # 6) Заполняем строки (даже если список пустой — это нормально)
    # Ограничиваем выдачу согласно top_n (максимум 500 каналов)
    
    # Сначала фильтруем список, исключая исходный канал по ID
    source_channel_id_int = int(source_channel.id)
    filtered_similar_list = []
    for item in similar_list:
        ch_id = item.get("channel_id")
        if ch_id is not None:
            ch_id_int = int(ch_id) if not isinstance(ch_id, int) else ch_id
            # Исключаем исходный канал из списка похожих
            if ch_id_int == source_channel_id_int:
                continue
        filtered_similar_list.append(item)
    
    # ОТНОСИТЕЛЬНАЯ нормализация: топ-1 = 100%, остальные относительно него
    max_score = 0.0
    if filtered_similar_list:
        scores = [float(item.get("score", 0.0)) for item in filtered_similar_list]
        max_score = max(scores) if scores else 1.0
    
    rows_added = 0
    rows_skipped_no_channel = 0
    row_number = 0  # Счетчик для нумерации
    
    # Получаем данные исходного канала для дополнительной проверки
    source_username = source_channel.username or ""
    source_title = getattr(source_channel, "title", "") or ""
    
    # Ограничиваем количество каналов согласно top_n
    for item in filtered_similar_list[:top_n]:
        ch_id = item.get("channel_id")
        
        # Убеждаемся, что ch_id - это int
        if ch_id is not None:
            ch_id = int(ch_id) if not isinstance(ch_id, int) else ch_id
        
        score = float(item.get("score", 0.0))
        ch = channels_map.get(ch_id)

        if not ch:
            rows_skipped_no_channel += 1
            continue

        # Дополнительная проверка: исключаем исходный канал по ID, username, title и subscribers
        ch_username = getattr(ch, "username", "") or ""
        ch_title = getattr(ch, "title", "") or ""
        ch_subscribers = getattr(ch, "subscribers", None)
        source_subscribers = getattr(source_channel, "subscribers", None)
        
        # Проверяем по ID (основная проверка)
        if ch.id == source_channel.id:
            continue
        
        # Дополнительная проверка по username (если оба не пустые)
        if source_username and ch_username and ch_username == source_username:
            continue
        
        # Дополнительная проверка по title и subscribers (если оба не пустые и совпадают)
        # Это помогает отфильтровать исходный канал даже если ID не совпадает
        if source_title and ch_title and ch_title.strip() == source_title.strip():
            if source_subscribers is not None and ch_subscribers is not None:
                if source_subscribers == ch_subscribers:
                    continue
            # Если subscribers не совпадают, но title совпадает - все равно пропускаем для безопасности
            continue

        # ГИБРИДНАЯ нормализация:
        # Если топ-канал имеет хороший score (>=0.5) → относительная (топ=100%)
        # Если топ-канал имеет низкий score (<0.5) → абсолютная с усилением
        if max_score >= 0.5:
            # Хорошая схожесть - используем относительную нормализацию
            normalized_score = score / max_score
        else:
            # Низкая схожесть - показываем реальные проценты с корректировкой
            # Используем sqrt для выравнивания (0.25 → 50%, 0.16 → 40%, 0.09 → 30%)
            normalized_score = math.sqrt(score)
        
        relevance_percent = round(normalized_score * 100, 1)
        
        # Преобразуем процент в текстовое описание степени схожести
        if relevance_percent >= 80:
            similarity_text = "Очень высокая"
        elif relevance_percent >= 60:
            similarity_text = "Высокая"
        elif relevance_percent >= 40:
            similarity_text = "Средняя"
        else:
            similarity_text = "Низкая"
        
        # Для ID-based каналов (username начинается с "id:") не создаём ссылку
        if ch.username and ch.username.startswith("id:"):
            link = ""  # Приватные каналы не имеют публичной ссылки
        else:
            link = f"https://t.me/{ch.username}" if ch.username else ""

        row_number += 1
        ws.append([
            row_number,
            similarity_text,
            link,
            getattr(ch, "subscribers", None),
            getattr(ch, "title", None),
            getattr(ch, "description", None),
        ])
        rows_added += 1

    # Применяем стили к данным
    if rows_added > 0:
        _apply_data_style(ws, start_row=2)
    
    # Автоподбор ширины колонок
    _auto_width(ws)
    
    # Замораживаем первую строку (заголовки)
    ws.freeze_panes = "A2"
    
    # Включаем автофильтр
    if rows_added > 0:
        ws.auto_filter.ref = ws.dimensions

    # 7) Путь и сохранение
    if base_dir is None:
        base_dir = Path(__file__).resolve().parents[2]

    reports_dir = base_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    # Формируем читаемое имя файла
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    channel_title = getattr(source_channel, "title", "") or ""
    channel_username = source_channel.username or f"id_{source_channel.id}"
    
    # Очищаем username от префикса id: и @
    if channel_username.startswith("id:"):
        channel_name = f"ID_{channel_username[3:]}"
    else:
        channel_name = channel_username.replace("@", "")
    
    # Используем название канала, если есть, иначе username
    if channel_title:
        safe_title = _sanitize_filename(channel_title, max_length=40)
        filename_base = f"Похожие_каналы_{safe_title}"
    else:
        safe_name = _sanitize_filename(channel_name, max_length=40)
        filename_base = f"Похожие_каналы_{safe_name}"
    
    filename = reports_dir / f"{filename_base}_{date_str}.xlsx"

    wb.save(filename)
    return filename
