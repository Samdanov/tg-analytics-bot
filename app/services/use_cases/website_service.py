# app/services/usecases/website_service.py

"""
Сервис для анализа веб-сайтов.
Аналогичен channel_service, но работает с сайтами.
"""

from pathlib import Path
from typing import Optional
from datetime import datetime

from app.services.website_parser.parser import parse_website
from app.services.llm.analyzer import analyze_text_content
from app.services.similarity_engine.website_similarity import find_similar_channels_by_keywords
from app.services.xlsx_generator import generate_similar_channels_xlsx
from app.core.logging import get_logger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

logger = get_logger(__name__)


def _auto_width(ws):
    """Автоматически подгоняет ширину колонок."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except Exception:
                value = ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 80)


def _sanitize_filename(text: str, max_length: int = 50) -> str:
    """Очищает текст для использования в имени файла."""
    text = re.sub(r'[<>:"/\\|?*]', '', text)
    text = text.replace(' ', '_')
    text = re.sub(r'_+', '_', text)
    if len(text) > max_length:
        text = text[:max_length]
    return text.strip('_')


def _apply_header_style(ws, header_row: int = 1):
    """Применяет стили к заголовкам таблицы."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="000000")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    ws.row_dimensions[header_row].height = 30


def _apply_data_style(ws, start_row: int = 2):
    """Применяет стили к данным таблицы."""
    border_side = Side(style="thin", color="CCCCCC")
    data_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    alignments = {
        1: Alignment(horizontal="center", vertical="center"),
        2: Alignment(horizontal="center", vertical="center"),
        3: Alignment(horizontal="left", vertical="center"),
        4: Alignment(horizontal="right", vertical="center"),
        5: Alignment(horizontal="left", vertical="top", wrap_text=True),
        6: Alignment(horizontal="left", vertical="top", wrap_text=True),
    }
    
    for row in ws.iter_rows(min_row=start_row):
        for idx, cell in enumerate(row, start=1):
            cell.border = data_border
            if idx in alignments:
                cell.alignment = alignments[idx]
            
            if idx == 2 and isinstance(cell.value, (int, float)):
                value = float(cell.value)
                if value >= 80:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value >= 60:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif value >= 40:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.number_format = "0.0"
            
            if idx == 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            
            if idx == 3 and cell.value and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.font = Font(color="0563C1", underline="single")
                cell.hyperlink = cell.value


async def generate_website_similar_channels_xlsx(
    url: str,
    similar_channels: list,
    keywords: list,
    base_dir: Optional[Path] = None
) -> Path:
    """
    Генерирует Excel файл с похожими каналами для сайта.
    
    Args:
        url: URL сайта
        similar_channels: Список похожих каналов
        keywords: Ключевые слова из анализа
        base_dir: Базовая директория для сохранения
    
    Returns:
        Путь к созданному файлу
    """
    if base_dir is None:
        base_dir = Path(__file__).resolve().parents[3]
    
    reports_dir = base_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Похожие каналы"
    
    # Заголовки таблицы
    headers = [
        "Дата создания",
        "Релевантность, %",
        "Ссылка на канал",
        "Подписчики",
        "Название канала",
        "Описание канала",
    ]
    ws.append(headers)
    
    # Применяем стили к заголовкам
    _apply_header_style(ws, header_row=1)
    
    created_str = datetime.utcnow().strftime("%d.%m.%Y")
    
    # Находим максимальный score для нормализации
    max_score = 0.0
    if similar_channels:
        scores = [ch.get("score", 0.0) for ch in similar_channels]
        max_score = max(scores) if scores else 1.0
    
    # Заполняем данные
    for ch_data in similar_channels:
        score = float(ch_data.get("score", 0.0))
        
        # Нормализуем score до 0-100%
        if max_score > 0:
            normalized_score = min(score / max_score, 1.0)
        else:
            normalized_score = min(score, 1.0)
        
        relevance_percent = round(normalized_score * 100, 1)
        
        username = ch_data.get("username", "")
        if username.startswith("id:"):
            link = ""
        else:
            link = f"https://t.me/{username}" if username else ""
        
        ws.append([
            created_str,
            relevance_percent,
            link,
            ch_data.get("subscribers"),
            ch_data.get("title"),
            "",  # Описание не доступно для сайтов
        ])
    
    # Применяем стили к данным
    if similar_channels:
        _apply_data_style(ws, start_row=2)
    
    # Автоподбор ширины колонок
    _auto_width(ws)
    
    # Замораживаем первую строку
    ws.freeze_panes = "A2"
    
    # Включаем автофильтр
    if similar_channels:
        ws.auto_filter.ref = ws.dimensions
    
    # Формируем читаемое имя файла
    date_str = datetime.utcnow().strftime("%Y-%m-%d")
    # Извлекаем домен из URL
    domain = url.replace("https://", "").replace("http://", "").split("/")[0]
    safe_domain = _sanitize_filename(domain, max_length=40)
    filename = reports_dir / f"Похожие_каналы_сайт_{safe_domain}_{date_str}.xlsx"
    
    wb.save(filename)
    
    logger.info(f"Generated Excel report: {filename}")
    return filename


async def run_website_analysis_pipeline(url: str, top_n: int = 10) -> tuple[Path, dict]:
    """
    Полный пайплайн анализа веб-сайта:
    1. Парсинг сайта
    2. Анализ через LLM (ключевые слова, ЦА)
    3. Поиск похожих каналов по ключевым словам
    4. Генерация Excel отчета
    
    Args:
        url: URL сайта
        top_n: Количество похожих каналов
    
    Returns:
        Путь к Excel файлу
    """
    logger.info(f"[WEBSITE] Начало анализа сайта: {url}")
    
    # 1. Парсинг сайта
    text, error = await parse_website(url)
    if error:
        logger.error(f"[WEBSITE] Ошибка парсинга: {error}")
        raise ValueError(f"Не удалось распарсить сайт: {error}")
    
    if not text:
        raise ValueError("Не удалось извлечь текст с сайта")
    
    logger.info(f"[WEBSITE] Извлечено текста: {len(text)} символов")
    
    # 2. Анализ через LLM
    logger.info(f"[WEBSITE] Запуск LLM анализа...")
    analysis_result = await analyze_text_content(
        text=text,
        title=url,
        description=""
    )
    
    keywords = analysis_result.get("keywords", [])
    audience = analysis_result.get("audience", "")
    logger.info(
        f"[WEBSITE] Анализ завершен. "
        f"Ключевые слова: {len(keywords)}, "
        f"ЦА: {len(audience)} символов, "
        f"Ключевые слова: {keywords[:5]}..."
    )
    
    if not keywords:
        raise ValueError("Не удалось извлечь ключевые слова из сайта")
    
    # 3. Поиск похожих каналов по ключевым словам
    similar_channels = await find_similar_channels_by_keywords(
        keywords=keywords,
        top_n=top_n
    )
    
    if not similar_channels:
        raise ValueError("Не найдено похожих каналов по ключевым словам")
    
    logger.info(f"[WEBSITE] Найдено похожих каналов: {len(similar_channels)}")
    
    # 4. Генерация Excel отчета
    report_path = await generate_website_similar_channels_xlsx(
        url=url,
        similar_channels=similar_channels,
        keywords=keywords
    )
    
    logger.info(f"[WEBSITE] Pipeline завершен. Отчет: {report_path}")
    return report_path, analysis_result
